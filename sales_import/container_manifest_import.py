"""Parse container manifest CSV / Excel (MARKS, SHOP#, ITEM NO., T.CTN, T.QTY, …).

Distinct from packing-list CSV: headers include MARKS 唛头, SHOP#, T.CTN, T.QTY, etc.
"""

from __future__ import annotations

import csv
import math
import os
import re
from pathlib import Path
from typing import Any, Callable

from sales_import.manifest_sections import (
    ManifestSection,
    append_section_label,
    append_manifest_payment,
    classify_section_banner,
    extract_yellow_subtotal_metrics,
    finalize_footer_totals,
    is_before_goods_header,
    is_document_footer_row,
    is_goods_left_header,
    is_likely_yellow_subtotal_row,
    looks_like_section_banner,
    merge_footer_from_line,
    should_persist_section_label,
)

_RE_CTNS = re.compile(r"([\d.,]+)\s*ctns?", re.I)
_RE_PCS = re.compile(r"([\d.,]+)\s*pcs?", re.I)
_RE_QPC = re.compile(r"([\d.,]+)\s*pcs?\s*/\s*ctn", re.I)
_RE_CBM = re.compile(r"([\d.,]+)\s*cbm", re.I)
_RE_KGS = re.compile(r"([\d.,]+)\s*kgs?", re.I)
_RE_RMB = re.compile(r"[¥￥]\s*([\d,]+(?:\.\d+)?)", re.I)
_RE_NUM = re.compile(r"-?[\d,]+(?:\.\d+)?")


def _freight_package_like(description: str | None) -> bool:
    """True for freight / courier / 'sending package' style descriptions (for optional CTN infer)."""
    if not description:
        return False
    u = description.lower()
    needles = (
        "dhl",
        "fedex",
        "ups",
        "freight",
        "sending package",
        "快递",
        "货运",
        "shipping fee",
        "物流",
    )
    return any(x in u for x in needles)


def _norm_cell(s: Any) -> str:
    return re.sub(r"\s+", " ", str(s or "").strip().lower())


def _cell_float(cell: Any) -> float | None:
    if cell is None or cell == "":
        return None
    if isinstance(cell, bool):
        return None
    if isinstance(cell, (int, float)):
        v = float(cell)
        return None if abs(v) < 1e-12 and v < 0 else v
    s = str(cell).strip()
    if not s:
        return None
    m = _RE_RMB.search(s)
    if m:
        try:
            return float(m.group(1).replace(",", ""))
        except ValueError:
            pass
    m = _RE_CBM.search(s)
    if m:
        try:
            return float(m.group(1).replace(",", ""))
        except ValueError:
            pass
    m = _RE_KGS.search(s)
    if m:
        try:
            return float(m.group(1).replace(",", ""))
        except ValueError:
            pass
    m = _RE_CTNS.search(s)
    if m:
        try:
            return float(m.group(1).replace(",", ""))
        except ValueError:
            pass
    m = _RE_PCS.search(s)
    if m:
        try:
            return float(m.group(1).replace(",", ""))
        except ValueError:
            pass
    m = _RE_NUM.search(s.replace(",", ""))
    if m:
        try:
            return float(m.group(0).replace(",", ""))
        except ValueError:
            return None
    return None


def _parse_qty_per_carton_from_packing(packing: str) -> float | None:
    if not packing:
        return None
    m = _RE_QPC.search(packing)
    if not m:
        return None
    try:
        return float(m.group(1).replace(",", ""))
    except ValueError:
        return None


def is_container_manifest_header_row(cells: list[str]) -> bool:
    if len(cells) < 10:
        return False
    joined = " ".join(_norm_cell(c) for c in cells[:16])
    if "marks" not in joined and "唛头" not in str(cells[0] or ""):
        return False
    if "t.ctn" not in joined and "t ctn" not in joined:
        return False
    if "t.qty" not in joined and "t qty" not in joined:
        return False
    return True


def _header_indices(headers: list[str]) -> dict[str, int]:
    """Map logical names to column indices from one header row."""
    norm = [_norm_cell(h) for h in headers]
    out: dict[str, int] = {}

    def find(pred: Callable[[str], bool]) -> int | None:
        for i, h in enumerate(norm):
            if pred(h):
                return i
        return None

    i = find(lambda h: "marks" in h or "唛头" in h)
    if i is not None:
        out["marks"] = i
    i = find(lambda h: "shop" in h and "#" in h.replace(" ", ""))
    if i is None:
        i = find(lambda h: h.startswith("shop"))
    if i is not None:
        out["shop"] = i
    i = find(lambda h: "item" in h and "no" in h)
    if i is not None:
        out["item_no"] = i
    i = find(lambda h: "description" in h and "goods" in h)
    if i is not None:
        out["desc_start"] = i
    i = find(lambda h: h == "packing" or h.startswith("packing"))
    if i is not None:
        out["packing"] = i
    i = find(lambda h: "t.ctn" in h.replace(" ", "") or h == "t ctn")
    if i is not None:
        out["t_ctn"] = i
    i = find(lambda h: "t.qty" in h.replace(" ", "") or h == "t qty")
    if i is not None:
        out["t_qty"] = i
    for key, pred in (
        ("h", lambda h: h == "h"),
        ("w", lambda h: h == "w"),
        ("l", lambda h: h == "l"),
    ):
        j = find(pred)
        if j is not None:
            out[key] = j
    i = find(lambda h: "unit" in h and "cbm" in h)
    if i is not None:
        out["unit_cbm"] = i
    i = find(lambda h: "t.cbm" in h.replace(" ", "") or h == "t cbm")
    if i is not None:
        out["t_cbm"] = i
    i = find(lambda h: "unit" in h and "weight" in h)
    if i is not None:
        out["unit_weight"] = i
    i = find(lambda h: "t.weight" in h.replace(" ", "") or h == "t weight")
    if i is not None:
        out["t_weight"] = i
    i = find(
        lambda h: "u.price" in h.replace(" ", "")
        or ("price" in h and "rmb" in h)
    )
    if i is not None:
        out["u_price"] = i
    i = find(lambda h: "t.amount" in h.replace(" ", "") or ("amount" in h and "rmb" in h))
    if i is not None:
        out["t_amount"] = i
    i = find(lambda h: "box" in h)
    if i is not None:
        out["box_no"] = i
    i = find(lambda h: "remmark" in h or "remark" in h)
    if i is not None:
        out["remark"] = i

    if "desc_start" in out and "packing" in out:
        ds = out["desc_start"]
        pk = out["packing"]
        if pk > ds + 1:
            out["desc_end"] = pk - 1
        else:
            out["desc_end"] = min(ds + 2, len(headers) - 1)
    return out


def _joined_row(r: list[str]) -> str:
    parts = [re.sub(r"\s+", " ", str(c or "").strip()) for c in r]
    return " ".join(p for p in parts if p).strip()


def _is_lone_text_row(r: list[str]) -> bool:
    """True when exactly one cell in the row has content.

    Section banners and stray notes in these exports are written as a single populated
    cell with every other column blank — a real product/footer/subtotal row always
    populates at least two columns. This catches heading wording the classifier has
    never seen (and any stray cell) without needing to know its exact text.
    """
    nonblank = [str(c).strip() for c in r if c and str(c).strip()]
    return len(nonblank) == 1


def _should_skip_misc_row(joined: str) -> bool:
    """Rows that are not banners, footers, subtotals, or products (edge noise)."""
    u = joined.strip().upper()
    if not u:
        return False
    if u == "MARKS" or u.startswith("MARKS 唛"):
        return True
    prefixes = ("PIVOC", "CREDIT SUPPORT", "YIWU-", "IF OUTSTANDING")
    if any(u.startswith(p) for p in prefixes):
        return True
    return False


def _subtotal_metrics_from_idx(row: list[str], idx: dict[str, int]) -> dict[str, float | None]:
    def g(name: str) -> str:
        j = idx.get(name)
        if j is None or j >= len(row):
            return ""
        return str(row[j] or "").strip()

    return {
        "total_cartons": _cell_float(g("t_ctn")),
        "total_cbm": _cell_float(g("t_cbm")),
        "total_weight_kg": _cell_float(g("t_weight")),
        "total_amount_rmb": _cell_float(g("t_amount")),
    }


def _merge_subtotal_metrics(
    from_joined: dict[str, float | None],
    from_idx: dict[str, float | None],
) -> dict[str, float | None]:
    out = dict(from_joined)
    for k, v in from_idx.items():
        if v is not None:
            out[k] = v
    return out


def _row_has_line_qty_pcs(row: list[str], idx: dict[str, int]) -> bool:
    """True when T.QTY holds a line quantity like ``156pcs`` (continuation rows must not be yellow subtotals)."""
    j = idx.get("t_qty")
    if j is None or j >= len(row):
        return False
    t = str(row[j] or "").strip()
    return bool(t and _RE_PCS.search(t))


def _process_manifest_data_rows(
    data_rows: list[list[str]],
    idx: dict[str, int],
    *,
    csv_line_base: int | None = None,
    trace_out: list[dict[str, Any]] | None = None,
) -> tuple[list[dict[str, Any]], dict[str, Any] | None, list[dict[str, Any]], list[dict[str, Any]]]:
    lines: list[dict[str, Any]] = []
    section_subtotals: list[dict[str, Any]] = []
    section_banners: list[dict[str, Any]] = []
    footer_acc: dict[str, Any] = {}
    footer_payments: list[dict[str, Any]] = []

    current_section: ManifestSection = "shipped"
    current_banner_label: str | None = None
    in_before_goods_block = False
    footer_locked = False
    last_marks = ""
    last_shop = ""

    def _trace(event: str, **extra: Any) -> None:
        if trace_out is None:
            return
        row: dict[str, Any] = {"event": event, **extra}
        trace_out.append(row)

    for j, r in enumerate(data_rows):
        csv_line = csv_line_base + j if csv_line_base is not None else None
        if not r or not any(str(c or "").strip() for c in r):
            _trace("skip_blank", csv_line=csv_line)
            continue
        if is_container_manifest_header_row(r):
            idx.clear()
            idx.update(_header_indices(r))
            _trace("header_repeat", csv_line=csv_line)
            continue

        joined = _joined_row(r)

        new_sec, banner_label = classify_section_banner(joined)
        if new_sec is not None:
            current_section = new_sec
            current_banner_label = banner_label
            if is_before_goods_header(joined):
                in_before_goods_block = True
            elif is_goods_left_header(joined):
                in_before_goods_block = False
            else:
                in_before_goods_block = False
            last_marks = ""
            last_shop = ""
            section_banners.append({"text": banner_label, "section": new_sec, "recognized": True})
            _trace(
                "section_banner",
                csv_line=csv_line,
                section=new_sec,
                banner_label=banner_label,
            )
            continue

        if is_document_footer_row(joined):
            footer_locked = True
            merge_footer_from_line(footer_acc, joined)
            append_manifest_payment(footer_payments, joined)
            _trace("footer_total_row", csv_line=csv_line)
            continue

        if footer_locked:
            # Real product/manifest rows never appear after the TOTAL row in these
            # exports — everything from here on is the trailing cost/payment summary
            # (REPACKING COST, GOODS BALANCE, payment terms, ...), however it's worded.
            # Capture it on footer_text instead of risking a fake product line.
            merge_footer_from_line(footer_acc, joined)
            append_manifest_payment(footer_payments, joined)
            _trace("footer_tail_row", csv_line=csv_line, text=joined)
            continue

        if _is_lone_text_row(r):
            # Heading-shaped text the classifier doesn't recognize yet -> its own
            # "other" bucket instead of guessing/carrying forward. A stray cell that
            # doesn't look like a heading (a bare number, a single word) leaves the
            # active section untouched -- never lose the text either way.
            if looks_like_section_banner(joined):
                current_section = "other"
                current_banner_label = joined
            section_banners.append({"text": joined, "section": current_section, "recognized": False})
            _trace("unclassified_section_banner", csv_line=csv_line, text=joined)
            continue

        is_yellow = (
            not _row_has_line_qty_pcs(r, idx)
            and (
                is_likely_yellow_subtotal_row(joined)
                or _is_aggregate_row(r, idx)
            )
        )
        if is_yellow:
            mj = extract_yellow_subtotal_metrics(joined)
            mi = _subtotal_metrics_from_idx(r, idx)
            metrics = _merge_subtotal_metrics(mj, mi)
            if not in_before_goods_block and any(v is not None for v in metrics.values()):
                section_subtotals.append({"section": current_section, **metrics})
                _trace(
                    "yellow_subtotal",
                    csv_line=csv_line,
                    section=current_section,
                    metrics=metrics,
                )
            else:
                _trace(
                    "yellow_subtotal_skipped",
                    csv_line=csv_line,
                    in_before_goods_block=in_before_goods_block,
                    metrics=metrics,
                )
            continue

        if _should_skip_misc_row(joined):
            _trace("skip_misc_banner", csv_line=csv_line)
            continue

        item = _row_to_manifest_line(r, idx, last_marks=last_marks, last_shop=last_shop)
        if not item:
            _trace("skip_no_item", csv_line=csv_line)
            continue

        item["section"] = current_section
        label_for_remark = (
            current_banner_label if should_persist_section_label(current_banner_label) else None
        )
        item["remarks"] = append_section_label(item.get("remarks"), label_for_remark)

        m = (item.get("marks") or "").strip()
        if m:
            last_marks = m
        s = (item.get("warehouse") or "").strip()
        if s:
            last_shop = s

        item_index = len(lines)
        lines.append(item)
        _trace(
            "item",
            csv_line=csv_line,
            item_index=item_index,
            marks=item.get("marks"),
            total_cartons=item.get("total_cartons"),
            section=item.get("section"),
        )

    footer_out = finalize_footer_totals(footer_acc)
    return lines, footer_out, section_subtotals, section_banners, footer_payments


def _is_aggregate_row(row: list[str], idx: dict[str, int]) -> bool:
    """Subtotal/correction row: no item identity, no per-line text — just leftover
    metrics (CTN and/or CBM/weight) left over from a merged cell in the source sheet.
    Doesn't require a CTN figure: some of these rows carry only a CBM/weight
    correction (e.g. "3.5016CBM 452.0KGS") with no carton count at all.
    """
    def g(name: str) -> str:
        j = idx.get(name)
        if j is None or j >= len(row):
            return ""
        return str(row[j] or "").strip()

    marks = g("marks")
    shop = g("shop")
    item_no = g("item_no")
    if marks or shop or item_no:
        return False
    t_qty = g("t_qty")
    if t_qty and _RE_PCS.search(t_qty):
        return False
    desc_bits = []
    ds = idx.get("desc_start")
    de = idx.get("desc_end", ds)
    if ds is not None:
        end = de if de is not None else ds + 2
        for j in range(ds, min(end + 1, len(row))):
            desc_bits.append(str(row[j] or "").strip())
    if any(desc_bits):
        return False
    t_ctn = g("t_ctn")
    has_ctn_signal = bool(t_ctn and _RE_CTNS.search(t_ctn))
    has_metric_signal = any(g(name) for name in ("t_cbm", "t_weight", "unit_cbm", "unit_weight"))
    return has_ctn_signal or has_metric_signal


def _row_to_manifest_line(
    row: list[str],
    idx: dict[str, int],
    *,
    last_marks: str,
    last_shop: str,
) -> dict[str, Any] | None:
    def g(name: str) -> str:
        j = idx.get(name)
        if j is None or j >= len(row):
            return ""
        return str(row[j] or "").strip()

    marks = g("marks") or last_marks
    shop = g("shop") or last_shop
    item_no = g("item_no")

    desc_parts: list[str] = []
    ds = idx.get("desc_start")
    de = idx.get("desc_end", ds)
    if ds is not None:
        end = de if de is not None else ds + 2
        for j in range(ds, min(end + 1, len(row))):
            t = str(row[j] or "").strip()
            if t:
                desc_parts.append(t)
    description = " | ".join(desc_parts) if desc_parts else ""

    packing = g("packing")
    t_ctn_s = g("t_ctn")
    t_qty_s = g("t_qty")

    if not marks and not description and not t_ctn_s and not t_qty_s:
        return None

    if not marks and not description:
        return None

    total_cartons = _cell_float(t_ctn_s)
    total_quantity = _cell_float(t_qty_s)
    unit_cbm = _cell_float(g("unit_cbm"))
    total_cbm = _cell_float(g("t_cbm"))
    unit_weight_kg = _cell_float(g("unit_weight"))
    total_weight_kg = _cell_float(g("t_weight"))
    unit_price_rmb = _cell_float(g("u_price"))
    total_amount_rmb = _cell_float(g("t_amount"))

    dim_h = _cell_float(g("h"))
    dim_w = _cell_float(g("w"))
    dim_l = _cell_float(g("l"))

    qpc = _parse_qty_per_carton_from_packing(packing)

    # Blank T.CTN but packing gives pcs/ctn — infer cartons from T.QTY (Excel often merges CTN visually).
    if (
        total_cartons is None
        and total_quantity is not None
        and total_quantity > 0
        and qpc is not None
        and qpc > 0
    ):
        total_cartons = float(max(1, math.ceil(float(total_quantity) / qpc - 1e-9)))

    # Freight / DHL / "sending package" row: PDF often has no T.CTN (not a physical carton line).
    # Default OFF so we do not assign 1 CTN. Set MANIFEST_INFER_FREIGHT_PACKAGE_ONE_CTN=1 to restore old behavior.
    _infer_freight = os.environ.get("MANIFEST_INFER_FREIGHT_PACKAGE_ONE_CTN", "0").strip().lower() in (
        "1",
        "true",
        "yes",
    )
    if (
        _infer_freight
        and total_cartons is None
        and total_amount_rmb is not None
        and float(total_amount_rmb) > 0
        and (total_quantity is None or float(total_quantity) == 0)
        and _freight_package_like(description)
    ):
        total_cartons = 1.0

    # Blank T.CTN and no pcs/ctn from packing: Excel merged CTN on large T.QTY lines (e.g. 100pcs).
    # Default OFF — enable MANIFEST_INFER_ONE_CTN_FROM_LARGE_QTY=1 if needed alongside freight rule.
    _infer_large = os.environ.get("MANIFEST_INFER_ONE_CTN_FROM_LARGE_QTY", "0").strip().lower() in (
        "1",
        "true",
        "yes",
    )
    _min_tq_large = float(os.environ.get("MANIFEST_INFER_ONE_CTN_MIN_TQTY", "50"))
    if (
        _infer_large
        and total_cartons is None
        and total_quantity is not None
        and float(total_quantity) >= _min_tq_large
    ):
        total_cartons = 1.0

    marks_stripped = marks.strip()
    # item_code: only when CSV has no MARKS cell (synthetic id); otherwise NULL — real MARKS → marks.
    if marks_stripped:
        item_code_val = None
    else:
        slug = re.sub(r"[^\w\-]+", "-", description)[:48].strip("-") or "line"
        base = (last_marks or "SUB").strip()
        item_code_val = f"{base}|{slug}"[:512]

    manifest_item_no: int | float | str | None = None
    if item_no:
        m = re.match(r"^([\d.]+)", item_no.replace(",", ""))
        if m:
            try:
                v = float(m.group(1))
                manifest_item_no = int(v) if abs(v - round(v)) < 0.01 else v
            except ValueError:
                manifest_item_no = item_no
        else:
            manifest_item_no = item_no.strip() or None

    remarks_parts = [g("remark"), g("box_no")]
    remarks = " | ".join(p for p in remarks_parts if p)

    raw: dict[str, Any] = {
        "marks": marks_stripped or None,
        "item_code": item_code_val,
        "description": description[:2000] if description else None,
        "delivery_no": None,
        "warehouse": shop.strip() or None,
        "shop": shop.strip() or None,
        "manifest_item_no": manifest_item_no,
        "packaging": packing[:500] if packing else None,
        "qty_per_carton": qpc,
        "total_cartons": total_cartons,
        "total_quantity": total_quantity,
        "unit_price_rmb": unit_price_rmb,
        "total_amount_rmb": total_amount_rmb,
        "dim_h_cm": dim_h,
        "dim_w_cm": dim_w,
        "dim_l_cm": dim_l,
        "unit_cbm": unit_cbm,
        "total_cbm": total_cbm,
        "unit_weight_kg": unit_weight_kg,
        "total_weight_kg": total_weight_kg,
        "remarks": remarks[:2000] if remarks else None,
        "_parse_source": "container_manifest",
        "_normalize_flags": ["manifest_import"],
        "model_source": "manifest",
    }
    return raw


def load_container_manifest_lines_from_csv(
    path: Path,
    *,
    encoding: str | None = None,
    trace_events: list[dict[str, Any]] | None = None,
) -> tuple[list[dict[str, Any]], dict[str, Any] | None, list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]]]:
    path = path.expanduser().resolve()
    if not path.is_file():
        raise FileNotFoundError(path)
    enc = (encoding or os.environ.get("CSV_ENCODING", "").strip() or "utf-8-sig")
    with path.open(encoding=enc, newline="") as f:
        rows = list(csv.reader(f))
    if not rows:
        return [], None, [], [], []

    header_idx: int | None = None
    idx: dict[str, int] | None = None
    for i, r in enumerate(rows):
        if len(r) < 10:
            continue
        if is_container_manifest_header_row(r):
            header_idx = i
            idx = _header_indices(r)
            break
    if header_idx is None or idx is None or "t_ctn" not in idx:
        raise ValueError(
            "Not a recognized container manifest CSV (expected headers: MARKS, T.CTN, T.QTY, …)."
        )

    trace_kw: dict[str, Any] = {}
    if trace_events is not None:
        trace_kw["csv_line_base"] = header_idx + 2  # 1-based file line of first data row
        trace_kw["trace_out"] = trace_events
    lines, footer_totals, section_subtotals, section_banners, footer_payments = _process_manifest_data_rows(
        rows[header_idx + 1 :], idx, **trace_kw
    )
    return lines, footer_totals, section_subtotals, section_banners, footer_payments


def load_container_manifest_lines_from_xlsx(
    path: Path,
    *,
    sheet_name: str | None = None,
) -> tuple[list[dict[str, Any]], dict[str, Any] | None, list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]]]:
    path = path.expanduser().resolve()
    if not path.is_file():
        raise FileNotFoundError(path)
    try:
        from openpyxl import load_workbook  # type: ignore[import-untyped]
    except ImportError as e:
        raise RuntimeError("Install openpyxl: pip install openpyxl") from e

    wb = load_workbook(path, read_only=True, data_only=True)
    sn = sheet_name or os.environ.get("EXCEL_SHEET", "").strip() or None
    if sn:
        if sn not in wb.sheetnames:
            wb.close()
            raise ValueError(f"Sheet {sn!r} not in workbook; available: {wb.sheetnames}")
        ws = wb[sn]
    else:
        ws = wb[wb.sheetnames[0]]

    rows_iter = ws.iter_rows(values_only=True)
    rows: list[list[str]] = []
    for tup in rows_iter:
        rows.append(["" if c is None else str(c) for c in tup])
    wb.close()

    if not rows:
        return [], None, [], [], []

    header_idx: int | None = None
    idx: dict[str, int] | None = None
    for i, r in enumerate(rows):
        if len(r) < 10:
            continue
        if is_container_manifest_header_row(r):
            header_idx = i
            idx = _header_indices(r)
            break
    if header_idx is None or idx is None or "t_ctn" not in idx:
        raise ValueError(
            "Not a recognized container manifest Excel (expected headers: MARKS, T.CTN, T.QTY, …)."
        )

    lines, footer_totals, section_subtotals, section_banners, footer_payments = _process_manifest_data_rows(
        rows[header_idx + 1 :], idx
    )
    return lines, footer_totals, section_subtotals, section_banners, footer_payments
