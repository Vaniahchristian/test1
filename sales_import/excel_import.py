"""Load packing-list rows from Excel (.xlsx) or CSV into the same dict shape as Reducto extract.

Expects a **header row** with columns similar to the delivery note PDF / CSV export.
Optional **TOTAL** row: first cell contains "TOTAL" / "合计"; numeric footer is read from that row.

Excel sheet: env **EXCEL_SHEET** (name) or pass sheet_name=...
CSV encoding: env **CSV_ENCODING** (default utf-8-sig).
"""

from __future__ import annotations

import csv
import logging
import os
import re
from pathlib import Path
from typing import Any

logger = logging.getLogger(__name__)

_OPTIONAL_FLOAT_FIELDS = frozenset(
    {
        "line_no",
        "total_cartons",
        "qty_per_carton",
        "total_quantity",
        "unit_price_rmb",
        "total_amount_rmb",
        "dim_l_cm",
        "dim_w_cm",
        "dim_h_cm",
        "unit_cbm",
        "total_cbm",
        "unit_weight_kg",
        "total_weight_kg",
    }
)

# (internal_key, substrings) — first substring match wins in column order for that key.
# More specific aliases must appear before looser ones in this list.
_HEADER_RULES: list[tuple[str, tuple[str, ...]]] = [
    # item_code before line_no: "ITEM NO." contains "no." so item_code must claim it first.
    ("item_code", ("item no", "item code", "产品货号", "item_no", "sku")),
    ("line_no", ("no.", "编号", "line no", "seq", "row no")),
    ("delivery_no", ("del no", "送货单号", "delivery")),
    ("customer_item_ref", ("cus no", "客户货号", "cust")),
    ("description", ("des.", "描述", "description", "品名英")),
    ("total_cartons", ("总箱数", "t.ctn", "ctn", "carton", "箱数")),
    ("qty_per_carton", ("每箱", "pcs/ctn", "qpc", "qty per", "per carton", "packing")),
    ("total_quantity", ("t.qty", "t qty", "总数量", "total qty", "ttl qty", "tq")),
    # unit_price_rmb before total_amount_rmb: "U.PRICE (RMB)" contains "rmb" which
    # total_amount_rmb would steal before unit_price_rmb gets a chance.
    ("unit_price_rmb", ("u/p", "u.price", "单价", "unit price")),
    ("total_amount_rmb", ("amount", "金额", "rmb", "amount rmb")),
    ("dim_l_cm", ("外箱长", "length", "l(cm)", "l (cm)", "dim l")),
    ("dim_w_cm", ("外箱宽", "width", "w(cm)", "w (cm)", "dim w")),
    ("dim_h_cm", ("外箱高", "height", "h(cm)", "h (cm)", "dim h")),
    ("unit_cbm", ("unit cbm", "cbm/ctn", "单箱体积")),
    ("total_cbm", ("ttl cbm", "total cbm", "line cbm", "t.t.cbm", "t.t. cbm", "t.cbm", "line vol")),
    # unit_weight_kg before unit: "UNIT WEIGHT" contains "unit" which the unit rule would steal.
    ("unit_weight_kg", ("unit weight", "g.w", "gw.", "unit kg", "单箱重量", "毛重")),
    (
        "total_weight_kg",
        ("t.weight", "ttl kgs", "ttl kg", "total kgs", "total kg", "总重量", "t.t.kgs", "t.t.kg", "t.t kg"),
    ),
    ("barcode", ("条形", "barcode", "ean")),
    ("warehouse", ("w.h", "warehouse", "仓库", "wh")),
    ("unit", ("unit", "单位")),
    ("product_name_local", ("品名", "name local", "产品名")),
    ("material", ("material", "材质")),
    ("remarks", ("rek", "remark", "备注")),
]

# Section-header rows found in some CSV exports; maps normalised first-cell text → section value.
_CSV_SECTION_MAP: dict[str, str] = {
    "new order": "shipped",
    "goods left behind": "left_in_warehouse",
    "repacked": "repacked",
}


def _norm_header(cell: Any) -> str:
    s = str(cell or "").strip().lower()
    s = re.sub(r"\s+", " ", s)
    return s


def _cell_to_float(cell: Any) -> float | None:
    if cell is None or cell == "":
        return None
    if isinstance(cell, bool):
        return None
    if isinstance(cell, (int, float)):
        v = float(cell)
        return None if v < 0 and abs(v) < 1e-9 else v
    s = str(cell).strip().replace(",", "")
    if not s or s in ("—", "-", "[ ]"):
        return None
    m = re.search(r"-?\d+(?:\.\d+)?", s.replace(" ", ""))
    if not m:
        return None
    try:
        return float(m.group(0))
    except ValueError:
        return None


def _cell_to_str(cell: Any) -> str | None:
    if cell is None:
        return None
    if isinstance(cell, float) and abs(cell - round(cell)) < 1e-9:
        s = str(int(round(cell)))
    else:
        s = str(cell).strip()
    return s or None


def _map_headers(headers: list[str]) -> dict[int, str]:
    """column_index (0-based) -> internal field name."""
    norm = [_norm_header(h) for h in headers]
    used_cols: set[int] = set()
    col_to_field: dict[int, str] = {}

    for key, aliases in _HEADER_RULES:
        for j, h in enumerate(norm):
            if j in used_cols or not h:
                continue
            if any(a in h for a in aliases):
                col_to_field[j] = key
                used_cols.add(j)
                break

    # Fallback: single "qty" column when no qty_per_carton — infer from context
    if not any(f == "qty_per_carton" for f in col_to_field.values()) and any(
        f == "total_quantity" for f in col_to_field.values()
    ):
        for j, h in enumerate(norm):
            if j in used_cols or not h:
                continue
            if h in ("qty", "quantity", "q") and "per" not in h:
                col_to_field[j] = "qty_per_carton"
                used_cols.add(j)
                break

    # Fallback: "l" / "w" / "h" single-letter headers (Excel often short)
    dim_short = {"l": "dim_l_cm", "w": "dim_w_cm", "h": "dim_h_cm"}
    for j, h in enumerate(norm):
        if j in used_cols or not h:
            continue
        if h in dim_short:
            field = dim_short[h]
            if not any(f == field for f in col_to_field.values()):
                col_to_field[j] = field
                used_cols.add(j)

    # PDF/CSV export: one header cell "L W H 外箱…" with L, W, H in the next two columns.
    for j, h in enumerate(norm):
        if j > len(norm) - 3:
            break
        if all(c in h for c in ("l", "w", "h")) and (
            "外箱" in h or "尺寸" in h or (" l " in f" {h} " and " w " in f" {h} " and " h " in f" {h} ")
        ):
            for k, name in enumerate(("dim_l_cm", "dim_w_cm", "dim_h_cm")):
                col_to_field[j + k] = name
            used_cols.update({j, j + 1, j + 2})
            break

    # Column whose header is literally "cbm …" but not "ttl cbm" (per-carton CBM).
    for j, h in enumerate(norm):
        if j in used_cols or not h:
            continue
        if h.startswith("cbm ") and "ttl" not in h:
            col_to_field[j] = "unit_cbm"
            used_cols.add(j)
            break

    return col_to_field


def _is_total_row(first_cell: Any) -> bool:
    s = _norm_header(first_cell)
    return "total" in s or "合计" in str(first_cell or "")


def _parse_footer_numbers(row_values: list[Any]) -> dict[str, Any] | None:
    nums: list[float] = []
    for c in row_values:
        v = _cell_to_float(c)
        if v is not None and v >= 0:
            nums.append(v)
    if len(nums) < 5:
        return None
    return {
        "total_cartons": nums[0],
        "total_quantity": nums[1],
        "total_amount_rmb": nums[2],
        "total_cbm": nums[3],
        "total_weight_kg": nums[4],
        "footer_text": " ".join(str(x) for x in row_values if x)[:800],
    }


def _row_to_line(
    row_values: list[Any],
    col_to_field: dict[int, str],
    *,
    source: str = "excel",
) -> dict[str, Any] | None:
    raw: dict[str, Any] = {}
    for j, field in col_to_field.items():
        if j >= len(row_values):
            continue
        cell = row_values[j]
        if field in _OPTIONAL_FLOAT_FIELDS:
            v = _cell_to_float(cell)
            if v is not None:
                if field == "line_no":
                    raw[field] = int(v) if abs(v - round(v)) < 0.01 else v
                else:
                    raw[field] = v
        else:
            s = _cell_to_str(cell)
            if s:
                raw[field] = s

    code = (raw.get("item_code") or "").strip() if raw.get("item_code") else ""
    if not code or code in ("[ ]",):
        return None
    raw["_parse_source"] = source
    flag = "csv_import" if source == "csv" else "excel_import"
    raw.setdefault("_normalize_flags", []).append(flag)
    return raw


def _find_csv_header_row_index(rows: list[list[str]]) -> int:
    """Locate the header row; column positions differ across exports (e.g. extra PHOTO cols)."""
    for i, r in enumerate(rows):
        if len(r) < 10:
            continue
        mapped = _map_headers(r)
        fields = set(mapped.values())
        if "item_code" in fields and "total_cartons" in fields:
            return i
    raise ValueError(
        "Could not find a packing-list header row (expect ITEM NO + CTN columns). "
        "Ensure the CSV has one header row like the PDF export."
    )


def load_sales_lines_from_csv(
    path: Path,
    *,
    encoding: str | None = None,
) -> tuple[list[dict[str, Any]], dict[str, Any] | None]:
    path = path.expanduser().resolve()
    if not path.is_file():
        raise FileNotFoundError(path)
    enc = (encoding or os.environ.get("CSV_ENCODING", "").strip() or "utf-8-sig")
    with path.open(encoding=enc, newline="") as f:
        rows = list(csv.reader(f))
    if not rows:
        return [], None
    hi = _find_csv_header_row_index(rows)
    headers = rows[hi]
    col_to_field = _map_headers(headers)
    if "item_code" not in col_to_field.values():
        raise ValueError(
            "CSV import could not map ITEM NO / 产品货号. Header row: "
            + repr(headers[:12])
        )

    # Detect description continuation columns: empty-header cols immediately after the
    # description col — arise when Excel exported a merged cell spanning several columns.
    desc_col = next((j for j, f in col_to_field.items() if f == "description"), None)
    desc_extra: list[int] = []
    if desc_col is not None:
        norm_hdrs = [_norm_header(h) for h in headers]
        j = desc_col + 1
        while j < len(headers) and j not in col_to_field and not norm_hdrs[j]:
            desc_extra.append(j)
            j += 1

    lines: list[dict[str, Any]] = []
    footer: dict[str, Any] | None = None
    current_section: str | None = None
    for r in rows[hi + 1 :]:
        if not r or all(v is None or str(v).strip() == "" for v in r):
            continue
        if r[0] and _is_total_row(r[0]):
            footer = _parse_footer_numbers(r)
            continue
        # Detect section-header rows (e.g. "NEW ORDER", "GOODS LEFT BEHIND").
        first_norm = _norm_header(r[0]) if r[0] else ""
        if first_norm in _CSV_SECTION_MAP:
            current_section = _CSV_SECTION_MAP[first_norm]
            continue
        item = _row_to_line(r, col_to_field, source="csv")
        if not item:
            continue
        # Skip non-product rows (financial notes, freight lines) that have no quantity data.
        if (
            item.get("total_cartons") is None
            and item.get("total_quantity") is None
            and not item.get("description")
        ):
            continue
        # Merge description from continuation columns when the mapped col was empty.
        if desc_extra and not item.get("description"):
            parts = [
                str(r[j]).strip().replace("\r\n", "\n").replace("\r", "\n")
                for j in desc_extra
                if j < len(r) and r[j] and str(r[j]).strip()
            ]
            if parts:
                item["description"] = "\n".join(parts)
        if current_section:
            item["section"] = current_section
        lines.append(item)

    return lines, footer


def load_sales_lines_from_xlsx(
    path: Path,
    *,
    sheet_name: str | None = None,
) -> tuple[list[dict[str, Any]], dict[str, Any] | None]:
    logger.info(f"Loading Excel file: {path}")
    path = path.expanduser().resolve()
    if not path.is_file():
        logger.error(f"File not found: {path}")
        raise FileNotFoundError(path)

    try:
        from openpyxl import load_workbook  # type: ignore[import-untyped]
    except ImportError as e:
        logger.error("openpyxl not installed")
        raise RuntimeError("Install openpyxl: pip install openpyxl") from e

    logger.info(f"Opening workbook: {path}")
    wb = load_workbook(path, read_only=True, data_only=True)
    sn = sheet_name or os.environ.get("EXCEL_SHEET", "").strip() or None
    logger.info(f"Available sheets: {wb.sheetnames}, selected: {sn or wb.sheetnames[0]}")
    if sn:
        if sn not in wb.sheetnames:
            wb.close()
            logger.error(f"Sheet {sn!r} not found in {wb.sheetnames}")
            raise ValueError(f"Sheet {sn!r} not in workbook; available: {wb.sheetnames}")
        ws = wb[sn]
    else:
        ws = wb[wb.sheetnames[0]]

    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        wb.close()
        logger.error("Excel file is empty (no rows)")
        return [], None

    headers = list(header_row)
    logger.info(f"Headers found: {headers[:15]}...")
    col_to_field = _map_headers(headers)
    logger.info(f"Mapped columns: {col_to_field}")
    if "item_code" not in col_to_field.values():
        wb.close()
        preview = [str(h) for h in headers[:30]]
        logger.error(f"Could not map item_code column. Headers: {preview}")
        raise ValueError(
            "Excel import could not find an item / SKU column. "
            "Name a column ITEM NO, item code, or 产品货号. "
            f"First row was: {preview}"
        )

    lines: list[dict[str, Any]] = []
    footer: dict[str, Any] | None = None
    row_count = 0

    for row in rows_iter:
        vals = list(row)
        row_count += 1
        if not vals or all(v is None or str(v).strip() == "" for v in vals):
            continue
        if _is_total_row(vals[0]):
            footer = _parse_footer_numbers(vals)
            logger.info(f"Footer/TOTAL row found: {footer}")
            continue
        item = _row_to_line(vals, col_to_field, source="excel")
        if item:
            lines.append(item)

    wb.close()
    logger.info(f"Excel parsing complete - rows scanned: {row_count}, lines parsed: {len(lines)}, footer: {footer is not None}")
    return lines, footer